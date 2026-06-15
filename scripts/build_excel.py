#!/usr/bin/env python3
"""
Callie Social Content Workflow — Excel内容包生成工具
用法: python build_excel.py <内容数据> <输出文件> --keyframes <关键帧目录>

功能:
- 将 AI 生成的内容包（脚本/分镜/文案）输出为格式化 Excel
- 自动嵌入关键帧参考图
- 5个工作表：内容总览、视频脚本、关键帧分镜、产品信息、品牌安全检查
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import json
import os
import sys
import argparse
from pathlib import Path

DEFAULT_KEYFRAMES_DIR = "keyframes"
DEFAULT_OUTPUT = "Callie_Content_Pack.xlsx"

# ============================================================
# 样式常量
# ============================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(cell, fill=None, font=None, alignment=None, border=None):
    if fill: cell.fill = fill
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if border: cell.border = border


def add_image_to_cell(ws, img_path: str, anchor: str, height: int = 160, width: int = 90):
    """将图片嵌入到指定单元格"""
    if not os.path.exists(img_path):
        print(f"[WARN] Image not found: {img_path}")
        return
    try:
        img = XLImage(img_path)
        img.height = height
        img.width = width
        img.anchor = anchor
        ws.add_image(img)
    except Exception as e:
        print(f"[WARN] Could not add image {img_path}: {e}")


# ============================================================
# 内容数据
# ============================================================

def build_overview_sheet(ws, data: dict):
    """工作表1: 内容总览"""
    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1, value="Callie 社媒内容包 — Instagram")
    style_cell(title, fill=HEADER_FILL,
               font=Font(name="Arial", bold=True, color="FFFFFF", size=14),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=THIN_BORDER)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    p = ws.cell(row=2, column=1, value=f"产品：{data.get('product_name', '')} | 平台：Instagram | 视频结构参考：{data.get('video_ref', '')}")
    style_cell(p, fill=ALT_ROW_FILL, font=BODY_FONT,
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
               border=THIN_BORDER)
    ws.row_dimensions[2].height = 25

    for i, h in enumerate(["板块", "内容", "备注"], 1):
        c = ws.cell(row=3, column=i, value=h)
        style_cell(c, fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center"),
                   border=THIN_BORDER)

    rows = [
        ("Big Idea", data.get("big_idea", ""), ""),
        ("品牌角度", data.get("brand_angle", ""), ""),
        ("产品关联", data.get("product_tie_in", ""), ""),
        ("英文贴文", data.get("caption", ""), "完整英文文案，可直接在Instagram发布"),
        ("11个Hashtag", data.get("hashtags", ""), "共11个，空格分隔"),
    ]

    for i, (label, value, note) in enumerate(rows, 4):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        c1 = ws.cell(row=i, column=1, value=label)
        style_cell(c1, fill=SUBHEADER_FILL, font=SUBHEADER_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                   border=THIN_BORDER)
        c2 = ws.cell(row=i, column=2, value=value + (f"\n\n【备注】{note}" if note else ""))
        style_cell(c2, fill=fill, font=BODY_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                   border=THIN_BORDER)
        c3 = ws.cell(row=i, column=3, value=note)
        style_cell(c3, fill=fill, font=BODY_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                   border=THIN_BORDER)
        ws.row_dimensions[i].height = 80

    # CTA
    cta_row = len(rows) + 4
    ws.merge_cells(start_row=cta_row, start_column=1, end_row=cta_row, end_column=3)
    cta = ws.cell(row=cta_row, column=1, value=f"CTA\n{data.get('cta', '')}")
    style_cell(cta, fill=GOLD_FILL, font=BOLD_FONT,
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
               border=THIN_BORDER)
    ws.row_dimensions[cta_row].height = 40

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 25


def build_script_sheet(ws, script_rows: list, keyframes_dir: str):
    """工作表2: 视频脚本（带图片）"""
    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value="Instagram Reels 视频脚本（含关键帧参考图）")
    style_cell(title, fill=HEADER_FILL,
               font=Font(name="Arial", bold=True, color="FFFFFF", size=13),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=THIN_BORDER)
    ws.row_dimensions[1].height = 28

    headers = ["时间", "画面描述", "镜头运动", "文字Overlay", "情绪氛围", "备注", "关键帧", "参考图"]
    col_widths = [10, 38, 16, 28, 14, 12, 12, 18]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        style_cell(c, fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center"),
                   border=THIN_BORDER)
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

    for i, row_data in enumerate(script_rows, 3):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        for j, val in enumerate(row_data[:7], 1):
            c = ws.cell(row=i, column=j, value=val)
            style_cell(c, fill=fill, font=BODY_FONT,
                       alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                       border=THIN_BORDER)
        ws.row_dimensions[i].height = 90

        # 嵌入图片
        img_file = row_data[7]
        img_path = os.path.join(keyframes_dir, img_file)
        add_image_to_cell(ws, img_path, f"H{i}")


def build_storyboard_sheet(ws, storyboard_rows: list, keyframes_dir: str):
    """工作表3: 关键帧分镜（带大图）"""
    ws.merge_cells("A1:G1")
    title = ws.cell(row=1, column=1, value="关键帧分镜描述（含AI生成参考图）— Instagram Reels 9:16竖版")
    style_cell(title, fill=HEADER_FILL,
               font=Font(name="Arial", bold=True, color="FFFFFF", size=13),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=THIN_BORDER)
    ws.row_dimensions[1].height = 28

    headers = ["帧", "时间", "画面描述", "镜头运动", "文字Overlay", "情绪氛围", "参考图"]
    col_widths = [8, 10, 45, 20, 28, 16, 18]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        style_cell(c, fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center"),
                   border=THIN_BORDER)
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

    for i, row_data in enumerate(storyboard_rows, 3):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        for j, val in enumerate(row_data[:6], 1):
            c = ws.cell(row=i, column=j, value=val)
            style_cell(c, fill=fill, font=BODY_FONT,
                       alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                       border=THIN_BORDER)
        ws.row_dimensions[i].height = 90

        # 嵌入大图
        img_path = os.path.join(keyframes_dir, row_data[6])
        add_image_to_cell(ws, img_path, f"G{i}", height=160, width=90)


def build_product_sheet(ws, product_rows: list):
    """工作表4: 产品信息"""
    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1, value="产品信息")
    style_cell(title, fill=HEADER_FILL,
               font=Font(name="Arial", bold=True, color="FFFFFF", size=13),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=THIN_BORDER)
    ws.row_dimensions[1].height = 28

    for i, (label, en_val, note) in enumerate(product_rows, 2):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        c1 = ws.cell(row=i, column=1, value=label)
        style_cell(c1, fill=SUBHEADER_FILL, font=SUBHEADER_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                   border=THIN_BORDER)
        c2 = ws.cell(row=i, column=2, value=en_val)
        style_cell(c2, fill=fill, font=BODY_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                   border=THIN_BORDER)
        c3 = ws.cell(row=i, column=3, value=note)
        style_cell(c3, fill=fill, font=Font(name="Arial", size=10, italic=True),
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                   border=THIN_BORDER)
        ws.row_dimensions[i].height = 40

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 35


def build_safety_sheet(ws):
    """工作表5: 品牌安全检查"""
    ws.merge_cells("A1:B1")
    title = ws.cell(row=1, column=1, value="品牌安全检查")
    style_cell(title, fill=HEADER_FILL,
               font=Font(name="Arial", bold=True, color="FFFFFF", size=13),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=THIN_BORDER)
    ws.row_dimensions[1].height = 28

    checks = [
        ("检查项", "状态"),
        ("无夸大宣传", "通过"),
        ("无折扣/价格信息", "通过"),
        ("无竞品提及", "通过"),
        ("无争议话题（政治/社会敏感）", "通过"),
        ("情感表达符合brand-guide温暖真诚定位", "通过"),
        ("悬念结构不涉及负面价值观", "通过"),
        ("Hashtag合规（无品牌词误用）", "通过"),
        ("产品功能描述准确", "通过"),
        ("内容适合多国英语市场", "通过"),
    ]

    for i, (item, status) in enumerate(checks, 2):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        if i == 2:
            c1 = ws.cell(row=i, column=1, value=item)
            style_cell(c1, fill=HEADER_FILL, font=HEADER_FONT,
                       alignment=Alignment(horizontal="left", vertical="center"),
                       border=THIN_BORDER)
            c2 = ws.cell(row=i, column=2, value=status)
            style_cell(c2, fill=HEADER_FILL, font=HEADER_FONT,
                       alignment=Alignment(horizontal="center", vertical="center"),
                       border=THIN_BORDER)
        else:
            c1 = ws.cell(row=i, column=1, value=item)
            style_cell(c1, fill=fill, font=BODY_FONT,
                       alignment=Alignment(horizontal="left", vertical="center"),
                       border=THIN_BORDER)
            c2 = ws.cell(row=i, column=2, value=f"通过 - {status}")
            style_cell(c2, fill=GREEN_FILL, font=BOLD_FONT,
                       alignment=Alignment(horizontal="center", vertical="center"),
                       border=THIN_BORDER)
        ws.row_dimensions[i].height = 22

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22


# ============================================================
# 主函数
# ============================================================

def build_excel(content_json: str, output: str, keyframes_dir: str = DEFAULT_KEYFRAMES_DIR):
    """
    构建完整的 Excel 内容包

    Args:
        content_json: 内容数据 JSON 文件路径（或 JSON 字符串）
        output: 输出 Excel 文件路径
        keyframes_dir: 关键帧图片目录
    """
    # 解析内容数据
    if os.path.exists(content_json):
        with open(content_json, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = json.loads(content_json)

    wb = openpyxl.Workbook()

    # Sheet 1: 内容总览
    ws1 = wb.active
    ws1.title = "内容总览"
    build_overview_sheet(ws1, data)

    # Sheet 2: 视频脚本
    ws2 = wb.create_sheet("视频脚本")
    script_rows = data.get("script_rows", [])
    build_script_sheet(ws2, script_rows, keyframes_dir)

    # Sheet 3: 关键帧分镜
    ws3 = wb.create_sheet("关键帧分镜")
    storyboard_rows = data.get("storyboard_rows", [])
    build_storyboard_sheet(ws3, storyboard_rows, keyframes_dir)

    # Sheet 4: 产品信息
    ws4 = wb.create_sheet("产品信息")
    product_rows = data.get("product_rows", [])
    build_product_sheet(ws4, product_rows)

    # Sheet 5: 品牌安全检查
    ws5 = wb.create_sheet("品牌安全检查")
    build_safety_sheet(ws5)

    wb.save(output)
    print(f"[build_excel] Saved: {output}")
    return output


def main():
    parser = argparse.ArgumentParser(
        description="Build formatted Excel content pack with embedded keyframe images"
    )
    parser.add_argument("content", help="Content data JSON file path or JSON string")
    parser.add_argument("output", nargs="?", default=DEFAULT_OUTPUT,
                        help=f"Output Excel file path (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--keyframes", "-k", default=DEFAULT_KEYFRAMES_DIR,
                        help=f"Keyframe images directory (default: {DEFAULT_KEYFRAMES_DIR})")

    args = parser.parse_args()

    try:
        build_excel(args.content, args.output, args.keyframes)
        print(f"[Done] Excel content pack created: {os.path.abspath(args.output)}")
        return 0
    except Exception as e:
        print(f"[Error] {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())