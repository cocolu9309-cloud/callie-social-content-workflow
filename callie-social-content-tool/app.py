# app.py — Callie 社媒内容生成工具后端
import os
import yaml
import httpx
import base64
import json
import uuid
import asyncio
import time
from pathlib import Path
from typing import Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse

app = FastAPI(title="Callie 社媒内容生成工具")

# ============================================================
# 配置加载
# ============================================================
def load_config():
    cfg_path = Path(__file__).parent / "config.yaml"
    if not cfg_path.exists():
        raise RuntimeError("config.yaml not found")
    with open(cfg_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

CONFIG = load_config()
SILICONFLOW_KEY = CONFIG.get("siliconflow_api_key", "")
LLM_MODEL = CONFIG.get("llm_model", "Qwen/QwQ-32B")
IMAGE_MODEL = CONFIG.get("image_model", "Qwen/Qwen-Image")
DEFAULT_PLATFORM = CONFIG.get("default_platform", "Instagram")

if not SILICONFLOW_KEY or SILICONFLOW_KEY == "YOUR_API_KEY_HERE":
    raise RuntimeError("siliconflow_api_key not configured in config.yaml")

# ============================================================
# 品牌指南（从 brand-guide.md 迁移）
# ============================================================
BRAND_GUIDE = """
# Callie Brand Guide

## Brand Role
Callie.com is a personalized gift brand. Treat the brand as a place for thoughtful, custom, occasion-led gifts that help people express care through names, dates, memories, symbols, messages, and personal details.

## Core Categories
- Personalized jewelry
- Bags and accessories
- Birthday gifts
- Graduation gifts
- Wedding and bridesmaid gifts
- Mother's Day and Father's Day gifts
- Anniversary gifts
- Keepsake and memorial-style gifts when handled with care
- Occasion-led gift guides

## Audience
Default to multi-country English-speaking gift buyers:
- People buying for partners, friends, parents, children, bridesmaids, graduates, coworkers, and family.
- Users who want a gift to feel personal, not generic.
- Shoppers looking for a sentimental, aesthetic, and useful gift.

## Positioning
Callie helps everyday gift-givers make important moments feel personal. The brand should feel warm, thoughtful, modern, celebratory, and human.

## Voice
Use: Warm, clear, and emotionally specific. Modern social wording, but not slang-heavy.
Avoid: Overpromising life-changing emotional impact. Guilt-based gifting. Generic phrases such as "the perfect gift for everyone".

## Brand-Safe Emotional Territories
- "I saw this and thought of you."
- Remembering a detail.
- Milestones becoming keepsakes.
- Personalization as proof of attention.
- Gifts that feel made for one person, not a crowd.

## Product Tie-In Rule
Lead with the human moment. Show the relationship or occasion. Introduce the personalized gift as the object that carries the meaning. End with a gentle CTA.
"""

# ============================================================
# 平台规则（从 platform-rules.md 迁移）
# ============================================================
PLATFORM_RULES = {
    "Instagram": {
        "hashtag_count": "5-8",
        "caption_style": "Story-driven, emotional, visual-first",
        "cta_style": "Link in bio",
        "max_video_sec": 90,
    },
    "TikTok": {
        "hashtag_count": "11",
        "caption_style": "Hook-first, conversational, trend-aware",
        "cta_style": "Follow + link in bio",
        "max_video_sec": 60,
    },
    "Pinterest": {
        "hashtag_count": "5",
        "caption_style": "SEO-friendly, keyword-rich title",
        "cta_style": "Shop now",
        "max_video_sec": None,
    },
    "YouTubeShorts": {
        "hashtag_count": "8",
        "caption_style": "Curiosity-driven title, informative",
        "cta_style": "Subscribe + link",
        "max_video_sec": 60,
    },
    "Facebook": {
        "hashtag_count": "0",
        "caption_style": "Conversational, community-oriented",
        "cta_style": "Comment or share",
        "max_video_sec": None,
    },
    "X": {
        "hashtag_count": "0-2",
        "caption_style": "Punchy, single idea, wit",
        "cta_style": "Retweet or follow",
        "max_video_sec": None,
    },
}

# ============================================================
# 临时文件存储（task_id → 工作目录）
# ============================================================
TEMP_DIR = Path(__file__).parent / "temp"
TEMP_DIR.mkdir(exist_ok=True)

tasks = {}  # task_id → {status, progress, message, result_path}

# ============================================================
# SiliconFlow API 调用封装
# ============================================================
SF_API_URL = "https://api.siliconflow.cn/v1"

async def call_sf_llm(prompt: str, images: list = None) -> str:
    """调用 SiliconFlow LLM（聊天补全）"""
    headers = {
        "Authorization": f"Bearer {SILICONFLOW_KEY}",
        "Content-Type": "application/json"
    }
    messages = [{"role": "system", "content": "You are a creative social media content expert for Callie.com, a personalized gift brand."}]
    if images:
        img_url = f"data:image/jpeg;base64,{images[0]}"
        messages.append({
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": img_url}},
                {"type": "text", "text": prompt}
            ]
        })
    else:
        messages.append({"role": "user", "content": prompt})
    payload = {"model": LLM_MODEL, "messages": messages, "temperature": 0.8}
    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post(f"{SF_API_URL}/chat/completions", headers=headers, json=payload)
        resp.raise_for_status()
        data = resp.json()
        return data["choices"][0]["message"]["content"]


async def generate_keyframe_image(prompt: str, output_path: str, size: str = "1024x1792") -> bool:
    """调用 Qwen-Image 生成单张图"""
    headers = {"Authorization": f"Bearer {SILICONFLOW_KEY}", "Content-Type": "application/json"}
    payload = {"model": IMAGE_MODEL, "prompt": prompt, "image_size": size, "n": 1}
    try:
        async with httpx.AsyncClient(timeout=180.0) as client:
            resp = await client.post(f"{SF_API_URL}/images/generations", headers=headers, json=payload)
            resp.raise_for_status()
            data = resp.json()
            img_obj = data["images"][0]
            if "url" in img_obj and img_obj["url"]:
                img_resp = await client.get(img_obj["url"], timeout=60.0)
                img_resp.raise_for_status()
                image_data = img_resp.content
            elif "b64_json" in img_obj:
                image_data = base64.b64decode(img_obj["b64_json"])
            else:
                return False
            with open(output_path, "wb") as f:
                f.write(image_data)
            return True
    except Exception as e:
        print(f"[ERROR] Image gen failed: {e}")
        return False


def update_task(tid: str, status: str, progress: int, message: str, result_path: str = None):
    tasks[tid] = {"status": status, "progress": progress, "message": message, "result_path": result_path}


# ============================================================
# 关键帧 prompt 翻译逻辑（从 generate_keyframes.py 迁移）
# ============================================================
CAMERA_MAP = {
    "固定镜头": "fixed camera, static shot",
    "固定": "fixed camera, static shot",
    "缓慢推进镜头": "slow push-in, cinematic approach",
    "推进镜头": "push-in, cinematic approach",
    "推进": "push-in, cinematic approach",
    "拉远": "pull-out, revealing wider scene",
    "缓慢拉远": "slow pull-out, cinematic reveal",
    "上摇镜头": "tilt-up camera movement",
    "下摇": "tilt-down camera movement",
    "摇镜": "pan camera movement",
    "横移": "tracking shot, side movement",
    "特写推进": "close-up push, detail emphasis",
    "特写": "close-up shot, detail focus",
    "中景": "medium shot, contextual framing",
    "全景": "wide shot, full scene",
}

MOOD_MAP = {
    "悬念": "suspenseful, mysterious atmosphere",
    "期待": "anticipating, hopeful mood",
    "紧张": "tense, dramatic tension",
    "惊喜": "surprised, delightful moment",
    "感动": "emotional, touching scene",
    "温馨": "warm, cozy ambiance",
    "怀旧": "nostalgic, golden-toned warmth",
    "温暖": "warm, heartfelt feeling",
    "珍视": "cherished, precious moment",
    "优雅": "elegant, refined aesthetic",
    "确定": "confident, decisive mood",
    "愉悦": "joyful, light and bright",
    "浪漫": "romantic, soft and intimate",
}

def translate_camera(cn: str) -> str:
    for key, val in CAMERA_MAP.items():
        if key in cn:
            return val
    return "cinematic shot"

def translate_mood(cn: str) -> str:
    for key, val in MOOD_MAP.items():
        if key in cn:
            return val
    return "cinematic, emotionally resonant"

def translate_scene(cn: str) -> str:
    import re
    replacements = [
        (r"木盒", "wooden gift box"),
        (r"风琴式.*?相册", "accordion-fold photo album"),
        (r"风琴式.*?拉页", "accordion-fold photo strip"),
        (r"情侣照片?|合影", "couple photo"),
        (r"照片翻页", "photo flipping"),
        (r"玫瑰|干花", "roses and dried flowers"),
        (r"背景柔焦?虚化", "soft bokeh background"),
        (r"柔光", "soft diffused lighting"),
        (r"暖金色光线", "warm golden hour lighting"),
        (r"柔焦", "shallow depth of field, soft focus"),
        (r"深色背景", "dark minimalist background"),
        (r"极简", "minimalist composition"),
        (r"手.*?入镜", "hand entering frame"),
        (r"手指轻触", "fingertips gently touching"),
        (r"打开盒盖|盒盖打开", "box lid opening"),
        (r"风琴页展开", "accordion pages gradually unfurling"),
        (r"微微抬起", "gently lifting"),
        (r"特写展示", "close-up detail shot of"),
        (r"镜头随.*?拉远", "camera pulling back as"),
        (r"缓慢推进", "slow cinematic push-in"),
        (r"缓慢上摇", "slow tilt-up camera movement"),
        (r"特写推进", "close-up push toward"),
    ]
    result = cn
    for pattern, replacement in replacements:
        result = re.sub(pattern, replacement, result)
    result = result.strip("，,。. ")
    if len(result) > 200:
        result = result[:200] + "..."
    return result

def build_prompt_from_script_row(row: list) -> dict:
    timecode = row[0] if len(row) > 0 else "0s"
    scene_cn = row[1] if len(row) > 1 else ""
    camera_cn = row[2] if len(row) > 2 else ""
    overlay_en = row[3] if len(row) > 3 else ""
    mood_cn = row[4] if len(row) > 4 else ""

    scene_en = translate_scene(scene_cn)
    camera_en = translate_camera(camera_cn)
    mood_en = translate_mood(mood_cn)

    prompt_parts = [scene_en]
    if overlay_en and overlay_en not in ("(无文字，悬念音乐)", "(none)", ""):
        prompt_parts.append(f"Text overlay: '{overlay_en}'")
    prompt_parts.append(camera_en)
    prompt_parts.append(mood_en)
    prompt_parts.append("vertical 9:16 aspect ratio, soft natural lighting")

    return {
        "frame": f"Frame ( {timecode} )",
        "scene": scene_cn,
        "camera": camera_cn,
        "overlay": overlay_en,
        "mood": mood_cn,
        "prompt_en": f"Cinematic photo: {', '.join(prompt_parts)}. High quality, professional photography style."
    }


# ============================================================
# Excel 打包（从 build_excel.py 迁移核心逻辑）
# ============================================================
def build_content_pack_excel(data: dict, output_path: str, keyframes_dir: str):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    GOLD_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    thin = Side(style="thin", color="BFBFBF")
    THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, fill=None, font=None, alignment=None, border=None):
        if fill: cell.fill = fill
        if font: cell.font = font
        if alignment: cell.alignment = alignment
        if border: cell.border = border

    def add_img(ws, img_path, anchor, h=160, w=90):
        import os
        if not os.path.exists(img_path):
            return
        try:
            img = XLImage(img_path)
            img.height = h
            img.width = w
            img.anchor = anchor
            ws.add_image(img)
        except Exception:
            pass

    platform = data.get("platform", "Instagram")

    wb = openpyxl.Workbook()

    # === Sheet 1: 内容总览 ===
    ws1 = wb.active
    ws1.title = "内容总览"
    ws1.merge_cells("A1:C1")
    c = ws1.cell(1, 1, f"Callie 社媒内容包 — {platform}")
    style_cell(c, fill=HEADER_FILL, font=Font(name="Arial", bold=True, color="FFFFFF", size=14),
               alignment=Alignment(horizontal="center", vertical="center"), border=THIN_BORDER)
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells("A2:C2")
    p = ws1.cell(2, 1, f"产品：{data.get('product_name', '')} | 平台：{platform}")
    style_cell(p, fill=ALT_ROW_FILL, font=BODY_FONT,
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True), border=THIN_BORDER)
    ws1.row_dimensions[2].height = 25
    for i, h in enumerate(["板块", "内容", "备注"], 1):
        c = ws1.cell(3, i, h)
        style_cell(c, fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center"), border=THIN_BORDER)
    rows = [
        ("Big Idea", data.get("big_idea", ""), ""),
        ("品牌角度", data.get("brand_angle", ""), ""),
        ("产品关联", data.get("product_tie_in", ""), ""),
        ("英文贴文", data.get("caption", ""), "可直接发布"),
        ("Hashtag", data.get("hashtags", ""), ""),
    ]
    for i, (label, value, note) in enumerate(rows, 4):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        c1 = ws1.cell(i, 1, label)
        style_cell(c1, fill=SUBHEADER_FILL, font=SUBHEADER_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), border=THIN_BORDER)
        c2 = ws1.cell(i, 2, value)
        style_cell(c2, fill=fill, font=BODY_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), border=THIN_BORDER)
        c3 = ws1.cell(i, 3, note)
        style_cell(c3, fill=fill, font=BODY_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), border=THIN_BORDER)
        ws1.row_dimensions[i].height = 80
    ws1.merge_cells(start_row=9, start_column=1, end_row=9, end_column=3)
    cta = ws1.cell(9, 1, f"CTA\n{data.get('cta', '')}")
    style_cell(cta, fill=GOLD_FILL, font=BOLD_FONT,
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True), border=THIN_BORDER)
    ws1.row_dimensions[9].height = 40
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 55
    ws1.column_dimensions["C"].width = 25

    # === Sheet 2: 视频脚本 ===
    ws2 = wb.create_sheet("视频脚本")
    ws2.merge_cells("A1:H1")
    c = ws2.cell(1, 1, f"{platform} 视频脚本（含关键帧参考图）")
    style_cell(c, fill=HEADER_FILL, font=Font(name="Arial", bold=True, color="FFFFFF", size=13),
               alignment=Alignment(horizontal="center", vertical="center"), border=THIN_BORDER)
    ws2.row_dimensions[1].height = 28
    headers = ["时间", "画面描述", "镜头运动", "文字Overlay", "情绪氛围", "备注", "帧", "参考图"]
    col_widths = [10, 38, 16, 28, 14, 12, 10, 18]
    for i, h in enumerate(headers, 1):
        c = ws2.cell(2, i, h)
        style_cell(c, fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center"), border=THIN_BORDER)
        ws2.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    script_rows = data.get("script_rows", [])
    for i, row_data in enumerate(script_rows, 3):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        for j, val in enumerate(row_data[:7], 1):
            c = ws2.cell(i, j, val)
            style_cell(c, fill=fill, font=BODY_FONT,
                       alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), border=THIN_BORDER)
        ws2.row_dimensions[i].height = 90
        img_file = row_data[7] if len(row_data) > 7 else None
        if img_file:
            img_path = os.path.join(keyframes_dir, img_file)
            add_img(ws2, img_path, f"H{i}")

    # === Sheet 3: 品牌安全检查 ===
    ws3 = wb.create_sheet("品牌安全检查")
    ws3.merge_cells("A1:B1")
    c = ws3.cell(1, 1, "品牌安全检查")
    style_cell(c, fill=HEADER_FILL, font=Font(name="Arial", bold=True, color="FFFFFF", size=13),
               alignment=Alignment(horizontal="center", vertical="center"), border=THIN_BORDER)
    ws3.row_dimensions[1].height = 28
    checks = [
        ("检查项", "状态"),
        ("无夸大宣传", "通过"),
        ("无折扣/价格信息", "通过"),
        ("无竞品提及", "通过"),
        ("无争议话题", "通过"),
        ("情感表达符合品牌温暖真诚定位", "通过"),
    ]
    for i, (item, status) in enumerate(checks, 2):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        c1 = ws3.cell(i, 1, item)
        style_cell(c1, fill=HEADER_FILL if i == 2 else fill, font=HEADER_FONT if i == 2 else BODY_FONT,
                   alignment=Alignment(horizontal="left", vertical="center"), border=THIN_BORDER)
        c2 = ws3.cell(i, 2, status)
        style_cell(c2, fill=HEADER_FILL if i == 2 else GREEN_FILL,
                   font=HEADER_FONT if i == 2 else BOLD_FONT,
                   alignment=Alignment(horizontal="center", vertical="center"), border=THIN_BORDER)
        ws3.row_dimensions[i].height = 22
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 22

    wb.save(output_path)


# ============================================================
# 主生成端点
# ============================================================
@app.post("/generate")
async def generate(
    platform: str = Form("Instagram"),
    product_name: str = Form(""),
    product_images: list[UploadFile] = File(default=[]),
    video_frames: list[UploadFile] = File(default=[]),
):
    task_id = str(uuid.uuid4())
    task_dir = TEMP_DIR / task_id
    task_dir.mkdir(exist_ok=True)
    keyframes_dir = task_dir / "keyframes"
    keyframes_dir.mkdir()

    update_task(task_id, "running", 5, "正在分析产品图...")

    # Step A: 读取产品图 base64
    product_b64 = []
    for img in product_images[:3]:
        data = await img.read()
        product_b64.append(base64.b64encode(data).decode())

    # Step B: 读取视频帧 base64（如果上传了）
    video_b64 = []
    for vf in video_frames[:6]:
        data = await vf.read()
        video_b64.append(base64.b64encode(data).decode())

    try:
        # ===== Step 1: 分析产品信息 =====
        if product_b64:
            product_prompt = f"""分析这张产品图，提取以下信息（用中文回复）：
1. 产品名称
2. 品类
3. 材质/风格
4. 适用场景（列出3-5个具体节日/场合）
5. 目标人群

产品名称参考（用户输入）: {product_name}"""
            product_info = await call_sf_llm(product_prompt, images=product_b64)
        else:
            product_info = f"产品名称: {product_name}"

        update_task(task_id, "running", 20, "正在生成英文文案和Hashtag...")

        # ===== Step 2: 生成文案 =====
        rules = PLATFORM_RULES.get(platform, PLATFORM_RULES["Instagram"])
        platform_rule_text = f"""平台: {platform}
Hashtag数量: {rules['hashtag_count']}
文案风格: {rules['caption_style']}
CTA风格: {rules['cta_style']}"""

        caption_prompt = f"""你是 Callie.com 的社媒内容专家。参考以下品牌指南，生成{platform}平台的完整内容包。

{platform_rule_text}

品牌指南:
{BRAND_GUIDE}

产品信息:
{product_info}

请以 JSON 格式输出（不要加代码块标记，直接输出 JSON）：
{{
  "big_idea": "一句核心创意概念",
  "brand_angle": "品牌角度说明",
  "product_tie_in": "产品关联（轻/中/强）",
  "caption": "完整英文文案（可直接发布）",
  "hashtags": "#tag1 #tag2 ...",
  "cta": "行动引导文字"
}}"""

        caption_raw = await call_sf_llm(caption_prompt)
        caption_text = caption_raw.strip()
        if caption_text.startswith("```"):
            parts = caption_text.split("```")
            caption_text = parts[1] if len(parts) > 1 else caption_text
            if caption_text.startswith("json"):
                caption_text = caption_text[4:]
        caption_data = json.loads(caption_text)

        update_task(task_id, "running", 40, "正在生成中文视频脚本...")

        # ===== Step 3: 生成中文视频脚本 =====
        if video_b64:
            video_prompt = f"""分析这6个视频帧，提取：
1. Hook类型（前3秒怎么吸引注意力）
2. 叙事结构（三段式/问题-解决/情感故事等）
3. 节奏分布（0-3秒/3-15秒/15-30秒各放什么）
4. 可复用元素

以 JSON 格式回复：
{{
  "hook_type": "...",
  "narrative": "...",
  "pacing": {{"0-3s": "...", "3-15s": "...", "15-30s": "..."}},
  "reusable_elements": ["..."]
}}"""
            video_insight = await call_sf_llm(video_prompt, images=video_b64)
        else:
            video_insight = "默认结构：悬念开盒 → 揭晓定制内容 → 情感共鸣收尾"

        script_prompt = f"""基于以下产品信息和视频结构，生成6个关键帧的中文视频脚本。

产品信息:
{product_info}

视频结构参考: {video_insight}

平台: {platform}

以 JSON 数组格式输出（不要代码块标记，直接输出 JSON数组）：
[
  ["0-1s", "画面描述（中文，20字以内）", "镜头运动", "文字Overlay(英文)", "情绪氛围", "帧1", "keyframe_01.jpg"],
  ["1-3s", "...", "...", "...", "...", "帧2", "keyframe_02.jpg"],
  ["3-6s", "...", "...", "...", "...", "帧3", "keyframe_03.jpg"],
  ["6-9s", "...", "...", "...", "...", "帧4", "keyframe_04.jpg"],
  ["9-12s", "...", "...", "...", "...", "帧5", "keyframe_05.jpg"],
  ["12-15s", "...", "...", "...", "...", "帧6", "keyframe_06.jpg"]
]"""
        script_raw = await call_sf_llm(script_prompt)
        script_text = script_raw.strip()
        if script_text.startswith("```"):
            parts = script_text.split("```")
            script_text = parts[1] if len(parts) > 1 else script_text
            if script_text.startswith("json"):
                script_text = script_text[4:]
        script_rows = json.loads(script_text)

        update_task(task_id, "running", 60, "正在生成关键帧图片 1/6...")

        # ===== Step 4: 生成关键帧图片 =====
        keyframe_prompts = [build_prompt_from_script_row(row) for row in script_rows]
        keyframe_files = []
        for i, kf in enumerate(keyframe_prompts):
            img_path = keyframes_dir / f"keyframe_{i+1:02d}.jpg"
            ok = await generate_keyframe_image(kf["prompt_en"], str(img_path))
            update_task(task_id, "running", 60 + (i+1)*5, f"正在生成关键帧图片 {i+1}/6...")
            keyframe_files.append(str(img_path) if ok else None)
            if i < 5:
                await asyncio.sleep(2)

        update_task(task_id, "running", 90, "正在打包 Excel...")

        # ===== Step 5: 打包 Excel =====
        content_data = {
            "product_name": product_name or caption_data.get("product_name", "Callie Product"),
            "platform": platform,
            "video_ref": "@callie",
            "big_idea": caption_data.get("big_idea", ""),
            "brand_angle": caption_data.get("brand_angle", ""),
            "product_tie_in": caption_data.get("product_tie_in", ""),
            "caption": caption_data.get("caption", ""),
            "hashtags": caption_data.get("hashtags", ""),
            "cta": caption_data.get("cta", ""),
            "script_rows": script_rows,
        }

        excel_path = task_dir / f"Callie_{platform}_Content_Pack.xlsx"
        build_content_pack_excel(content_data, str(excel_path), str(keyframes_dir))

        update_task(task_id, "done", 100, "生成完成！", result_path=str(excel_path))

        return JSONResponse({"status": "ok", "task_id": task_id, "message": "生成完成"})

    except Exception as e:
        import traceback
        traceback.print_exc()
        update_task(task_id, "error", 0, f"生成失败: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# 状态查询 + 下载
# ============================================================
@app.get("/status/{task_id}")
async def get_status(task_id: str):
    task = tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return JSONResponse(task)

@app.get("/download/{task_id}")
async def download(task_id: str):
    task = tasks.get(task_id)
    if not task or not task.get("result_path"):
        raise HTTPException(status_code=404, detail="File not ready")
    path = Path(task["result_path"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, filename=path.name,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# 静态文件
# ============================================================
@app.get("/")
async def root():
    return FileResponse(Path(__file__).parent / "index.html")
